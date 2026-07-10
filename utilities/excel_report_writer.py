"""Write downloadable Excel test reports with expected vs actual results."""

from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utilities.config_reader import ConfigReader
from utilities.excel_sanitize import sanitize_dataframe, sanitize_excel_value
from utilities.logger import get_logger
from utilities.test_result_store import TestResultStore, RunTestSummary, ValidationRecord

logger = get_logger(__name__)

_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_MAX_COL_WIDTH = 60


def _status_fill(status: str):
    value = (status or "").upper()
    if value in ("PASS", "PASSED"):
        return _PASS_FILL
    if value in ("FAIL", "FAILED", "ERROR"):
        return _FAIL_FILL
    if value in ("WARNING", "WARN"):
        return _WARN_FILL
    return None


def _auto_width(ws, max_width=_MAX_COL_WIDTH):
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for col_idx in range(1, ws.max_column + 1):
        length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row, 500) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                length = max(length, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 12), max_width)


def _style_header_row(ws, row=1):
    if ws.max_row < row:
        return
    for cell in ws[row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _color_status_column(ws, status_col_name: str, header_row: list):
    if status_col_name not in header_row:
        return
    col_idx = header_row.index(status_col_name) + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        fill = _status_fill(str(cell.value or ""))
        if fill:
            cell.fill = fill


def _build_run_summary(meta: dict, summaries: list[RunTestSummary]) -> pd.DataFrame:
    passed = sum(1 for s in summaries if s.status == "PASSED")
    failed = sum(1 for s in summaries if s.status in ("FAILED", "ERROR"))
    skipped = sum(1 for s in summaries if s.status == "SKIPPED")
    rows = [
        {"Metric": "Run Date/Time", "Value": meta.get("run_started", datetime.now().isoformat())},
        {"Metric": "Environment", "Value": meta.get("environment", "N/A")},
        {"Metric": "Base URL", "Value": meta.get("base_url", "N/A")},
        {"Metric": "Total Tests", "Value": len(summaries)},
        {"Metric": "Passed", "Value": passed},
        {"Metric": "Failed", "Value": failed},
        {"Metric": "Skipped", "Value": skipped},
        {"Metric": "Pass Rate", "Value": f"{(passed / len(summaries) * 100):.1f}%" if summaries else "N/A"},
        {"Metric": "HTML Report", "Value": meta.get("html_report", "reports/report.html")},
        {"Metric": "Log File", "Value": meta.get("log_file", "N/A")},
    ]
    return pd.DataFrame(rows)


def _build_test_summary_df(summaries: list[RunTestSummary]) -> pd.DataFrame:
    if not summaries:
        return pd.DataFrame(
            columns=[
                "Test ID",
                "Test Name",
                "Module",
                "Partner",
                "Product",
                "Status",
                "Duration (sec)",
                "Failure Message",
                "Screenshot",
            ]
        )
    return pd.DataFrame(
        [
            {
                "Test ID": s.test_id,
                "Test Name": s.test_name,
                "Module": s.module,
                "Partner": s.partner,
                "Product": s.product,
                "Status": s.status,
                "Duration (sec)": s.duration_sec,
                "Failure Message": s.failure_message,
                "Screenshot": s.screenshot,
            }
            for s in summaries
        ]
    )


def _build_detail_df(validations: list[ValidationRecord]) -> pd.DataFrame:
    if not validations:
        return pd.DataFrame(
            columns=[
                "Test ID",
                "Test Name",
                "Partner",
                "Product",
                "Step #",
                "Step Description",
                "Validation Field",
                "Expected Result",
                "Actual Result",
                "Status",
                "Message",
                "Timestamp",
            ]
        )
    return pd.DataFrame(
        [
            {
                "Test ID": v.test_id,
                "Test Name": v.test_name,
                "Partner": v.partner,
                "Product": v.product,
                "Step #": v.step_num,
                "Step Description": v.step_title,
                "Validation Field": v.field_name,
                "Expected Result": v.expected,
                "Actual Result": v.actual,
                "Status": v.status,
                "Message": v.message,
                "Timestamp": v.timestamp,
            }
            for v in validations
        ]
    )


def _build_qa_recommendations(
    validations: list[ValidationRecord],
    summaries: list[RunTestSummary],
) -> pd.DataFrame:
    rows = [
        {
            "Area": "Test Data",
            "Finding": "Excel drives all Partner Spotlight validations",
            "Recommendation": "Review testdata/partner_products.xlsx after every site content change; keep enabled=FALSE for WIP rows.",
            "Priority": "High",
        },
        {
            "Area": "Reporting",
            "Finding": "HTML + Allure + Excel reports generated per run",
            "Recommendation": "Attach this Excel file to JIRA/tickets; use Detailed Validations sheet for defect steps.",
            "Priority": "High",
        },
        {
            "Area": "Execution",
            "Finding": "Headless runs are faster; visual mode helps demos",
            "Recommendation": "Use python run_tests.py visual for stakeholder walkthroughs; use default headless for CI/nightly.",
            "Priority": "Medium",
        },
        {
            "Area": "Coverage",
            "Finding": "26-step flow covers listing + detail + categories + PDF",
            "Recommendation": "Add negative tests (invalid partner, wrong product type) and boundary cases (empty search, geo filters).",
            "Priority": "Medium",
        },
        {
            "Area": "Stability",
            "Finding": "Menu navigation sometimes falls back to direct URL",
            "Recommendation": "Track menu-navigation failures separately; add retry or dedicated nav smoke test.",
            "Priority": "Medium",
        },
        {
            "Area": "Performance",
            "Finding": "Category filter application can take several minutes",
            "Recommendation": "Measure step-7 duration per product; set SLA thresholds and flag slow runs in Excel.",
            "Priority": "Low",
        },
        {
            "Area": "Cross-browser",
            "Finding": "Default browser is Chromium only",
            "Recommendation": "Run periodic Firefox/WebKit matrix on critical paths (PS-001, PS-002).",
            "Priority": "Low",
        },
    ]

    failed = [v for v in validations if v.status == "FAIL"]
    category_fails = [v for v in failed if v.field_name == "category_subcategory"]
    if category_fails:
        rows.insert(
            0,
            {
                "Area": "Category Validation",
                "Finding": f"{len(category_fails)} category/sub-category mismatch(es) in this run",
                "Recommendation": "Compare Excel ticket text with live sidebar labels and detail-page tags; update aliases in category_validator.py or fix site content.",
                "Priority": "Critical",
            },
        )

    feature_fails = [v for v in failed if str(v.field_name).startswith("Feature:")]
    if feature_fails:
        rows.insert(
            0,
            {
                "Area": "Features Validation",
                "Finding": f"{len(feature_fails)} feature(s) from expected_features missing on detail page",
                "Recommendation": "Update expected_features in Excel to match live Features section text exactly, or fix site content.",
                "Priority": "High",
            },
        )

    partner_fails = [v for v in failed if "partner" in v.field_name.lower()]
    if partner_fails:
        rows.insert(
            0,
            {
                "Area": "Partner Filter",
                "Finding": f"{len(partner_fails)} partner filter validation failure(s)",
                "Recommendation": "Verify partner_dropdown_label column matches exact dropdown text on Partner Spotlight.",
                "Priority": "High",
            },
        )

    search_fails = [v for v in failed if "search" in v.field_name.lower()]
    if search_fails:
        rows.insert(
            0,
            {
                "Area": "Search",
                "Finding": f"{len(search_fails)} search-related failure(s)",
                "Recommendation": "Confirm search_term in Excel matches catalog header count and unique product URLs.",
                "Priority": "High",
            },
        )

    if any(s.status in ("FAILED", "ERROR") for s in summaries):
        rows.insert(
            0,
            {
                "Area": "Defect Triage",
                "Finding": f"{sum(1 for s in summaries if s.status in ('FAILED', 'ERROR'))} test case(s) failed",
                "Recommendation": "Open bugs with Test ID, Step #, Expected vs Actual from Detailed Validations sheet; attach failure screenshot path.",
                "Priority": "Critical",
            },
        )

    warnings = [v for v in validations if v.status == "WARNING"]
    if warnings:
        rows.append(
            {
                "Area": "Optional Checks",
                "Finding": f"{len(warnings)} warning(s) (Quick View, Related Products, etc.)",
                "Recommendation": "Decide with product owner whether warnings should become hard failures or stay informational.",
                "Priority": "Medium",
            }
        )

    return pd.DataFrame(rows)


def _write_workbook(path: str, frames: dict[str, pd.DataFrame]) -> None:
    """Write sanitized sheets and apply styling."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            safe_df = sanitize_dataframe(df)
            safe_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            _style_header_row(ws)
            _auto_width(ws)
            headers = [sanitize_excel_value(c.value) for c in ws[1]]
            if sheet_name in ("Test Summary", "Detailed Validations"):
                _color_status_column(ws, "Status", headers)


def _validate_workbook(path: str) -> None:
    """Ensure workbook can be opened before publishing."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if not wb.sheetnames:
            raise ValueError("Excel report has no sheets")
    finally:
        wb.close()


def _publish_file(source: str, destination: str) -> None:
    """Atomically replace destination with a validated copy."""
    _validate_workbook(source)
    tmp_dest = f"{destination}.tmp.xlsx"
    try:
        if os.path.exists(tmp_dest):
            os.remove(tmp_dest)
        with open(source, "rb") as src, open(tmp_dest, "wb") as dst:
            dst.write(src.read())
        os.replace(tmp_dest, destination)
    except OSError:
        if os.path.exists(tmp_dest):
            try:
                os.remove(tmp_dest)
            except OSError:
                pass
        raise


class ExcelReportWriter:
    """Export TestResultStore data to a styled .xlsx file."""

    @staticmethod
    def write(output_path: str | None = None) -> str:
        reports_dir = ConfigReader.get_path("report_path", "reports")
        os.makedirs(reports_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            output_path = os.path.join(reports_dir, f"test_results_{timestamp}.xlsx")

        meta = TestResultStore.get_session_meta()
        summaries = TestResultStore.get_summaries()
        validations = TestResultStore.get_validations()

        frames = {
            "Run Summary": _build_run_summary(meta, summaries),
            "Test Summary": _build_test_summary_df(summaries),
            "Detailed Validations": _build_detail_df(validations),
            "QA Recommendations": _build_qa_recommendations(validations, summaries),
        }

        tmp_path = f"{output_path}.tmp.xlsx"
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            _write_workbook(tmp_path, frames)
            _validate_workbook(tmp_path)
            os.replace(tmp_path, output_path)
        except Exception:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
            raise
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

        latest = os.path.join(reports_dir, "latest_test_results.xlsx")
        try:
            _publish_file(output_path, latest)
        except OSError as exc:
            logger.warning("Could not update latest_test_results.xlsx: %s", exc)

        logger.info("Excel test report written: %s", output_path)
        return output_path
