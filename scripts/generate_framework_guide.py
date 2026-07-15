"""Generate step-by-step Excel documentation for the Intel Edge AI automation framework."""

from datetime import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

OUT_DIR = os.path.join(ROOT, "docs")
OUT_PATH = os.path.join(OUT_DIR, "Intel_Edge_AI_Automation_Framework_Guide.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FILL = PatternFill("solid", fgColor="D6E3F0")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, headers, rows, widths, row_height=40):
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = WRAP
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = ALT_FILL
        ws.row_dimensions[r].height = row_height
    autosize(ws, widths)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()

    # 1. Overview
    ws = wb.active
    ws.title = "1. Overview"
    overview = [
        ["Item", "Details"],
        ["Document Title", "Intel Edge AI Automation Framework — Step-by-Step Guide"],
        ["Project", "intel_edge_ai_automation"],
        [
            "Purpose",
            "Automate Partner Spotlight (Edge AI Catalog) validation end-to-end using "
            "Playwright + pytest. Excel-driven data. Produces HTML, Allure, and Excel "
            "reports with Expected vs Actual.",
        ],
        ["Application Under Test", "https://builders.intel.com (Edge AI Catalog / Partner Spotlight)"],
        ["Tech Stack", "Python 3.13, Playwright (sync), pytest, pytest-html, allure-pytest, pandas/openpyxl"],
        ["Design Pattern", "Page Object Model (POM) + Excel-driven data + Step Reporter"],
        ["Entry Point", "run_tests.py  (do NOT run conftest.py directly)"],
        ["Test Data File", "testdata/partner_products.xlsx  (sheet: PartnerProducts)"],
        ["Key Test File", "tests/test_partner_spotlight.py → PartnerSpotlight26StepFlow (26 steps)"],
        [
            "Catalog Smoke",
            "tests/test_catalog_features.py (CAT-001 metadata, CAT-002 view toggle after search, CAT-003 search)",
        ],
        ["Document Created", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    for r, row in enumerate(overview, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = WRAP
            cell.border = THIN
            if r == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif c == 1:
                cell.font = Font(bold=True)
                cell.fill = LABEL_FILL
        ws.row_dimensions[r].height = 24 if r == 1 else 42
    autosize(ws, [28, 100])

    # 2. How to Run
    ws2 = wb.create_sheet("2. How to Run")
    run_rows = [
        [1, "Create / activate venv", r".\venv\Scripts\Activate.ps1", "Project uses local venv"],
        [2, "Install dependencies", "pip install -r requirements.txt", "Includes playwright, pytest, openpyxl, pandas"],
        [3, "Install Chromium browser", "python -m playwright install chromium", "Required once (or after Playwright upgrade)"],
        [4, "Prepare test data", r"Edit testdata\partner_products.xlsx", "Set enabled=TRUE only for rows to run"],
        [5, "Run full suite (headless)", "python run_tests.py", "Default CI-friendly mode"],
        [6, "Run with visible browser", "python run_tests.py visual", "Headed + slow-mo 600ms"],
        [7, "Custom headed slow-mo", "python run_tests.py --headed --slow-mo 800", "For demos"],
        [8, "Smoke only", "python run_tests.py -m smoke", "Catalog + listing smoke"],
        [9, "Partner Spotlight only", "python run_tests.py tests/test_partner_spotlight.py", "26-step + listing smoke"],
        [10, "Single product (example)", "python run_tests.py tests/test_partner_spotlight.py -k PS-001", "Filter by test id"],
        [11, "Disable retries", "python run_tests.py --reruns 0", "Overrides default re-run behavior"],
        [12, "Open HTML report", r"reports\report.html", "Self-contained HTML"],
        [13, "Open Excel report", r"reports\latest_test_results.xlsx", "Always latest Expected vs Actual workbook"],
    ]
    write_table(
        ws2,
        ["#", "Action", "Command / Path", "Notes"],
        run_rows,
        [6, 32, 70, 40],
        row_height=28,
    )

    # 3. 26-Step Flow
    ws3 = wb.create_sheet("3. 26-Step Flow")
    steps = [
        [
            1,
            "Open Intel Builders home page",
            "Navigation",
            "Opens home URL from config",
            "Home page loads",
            "config base_url",
            "pages/home_page.py → open()",
            "Yes",
        ],
        [
            2,
            "Open Engagement menu",
            "Navigation",
            "Opens Engagement menu in header",
            "Engagement menu visible",
            "—",
            "pages/home_page.py → open_engagement_menu()",
            "Yes",
        ],
        [
            3,
            "Navigate to Edge AI Catalog",
            "Navigation",
            "Clicks menu path to Edge AI Catalog; falls back to direct URL if needed",
            "Edge AI Catalog page opens",
            "config edge_catalog_url",
            "home_page + edge_catalog_page",
            "Yes",
        ],
        [
            4,
            "Click Explore Partner Spotlight",
            "Navigation",
            "Clicks Explore Partner Spotlight (or opens Partner Spotlight URL)",
            "Partner Spotlight listing loads",
            "—",
            "edge_catalog_page + partner_spotlight_page",
            "Yes",
        ],
        [
            5,
            "Verify Edge AI Partner Spotlight page opens",
            "Listing",
            "Confirms URL contains partner-spotlight; catalog product count > 0",
            "Page loaded with products",
            "—",
            "partner_spotlight_page.verify_page_loaded()",
            "Yes",
        ],
        [
            6,
            "Filter listing by partner",
            "Listing / Filters",
            "Selects partner in Filter By Partners dropdown",
            "Partner filter applied",
            "partner_name, partner_dropdown_label",
            "PartnerSpotlightPage.select_partner()",
            "Yes",
        ],
        [
            7,
            "Verify product listing is displayed",
            "Listing / Filters",
            "Selects Application/System tab; applies category_subcategory sidebar filters; verifies product listed and product type badge",
            "Product visible with correct type badge",
            "product_type, category_subcategory, product_name",
            "PartnerSpotlight26StepFlow._ensure_product_type_and_listing()",
            "Yes",
        ],
        [
            8,
            "Validate application name on listing",
            "Listing",
            "Finds product name on listing card",
            "Application/product name matches Excel",
            "product_name",
            "listing_validator.validate_application_name()",
            "Yes",
        ],
        [
            9,
            "Validate short description on listing",
            "Listing",
            "Validates short description keywords on card",
            "Short description matches Excel snippet",
            "expected_short_description",
            "listing_validator.validate_short_description()",
            "Yes",
        ],
        [
            10,
            "Search product and verify results",
            "Search",
            "Searches by product/search_term (pSearch). Confirms catalog header count and product in results",
            "Catalog of N Unique Products; product found",
            "search_term / product_name",
            "PartnerSpotlightPage.search_products()",
            "Yes",
        ],
        [
            11,
            "Verify Grid View and List View on search results",
            "Search Results UI",
            "AFTER search: switches Grid then List; verifies active toggle + product still visible (matches client screenshots)",
            "Grid active with product; List active with product",
            "search_term, product_name",
            "validate_grid_view / validate_list_view",
            "Yes",
        ],
        [
            12,
            "Apply Category/Sub-Category filters (validate)",
            "Categories",
            "Collects listing site tags/text and compares each Excel category_subcategory pair (full strict check at Step 25)",
            "Pairs reported; product still listed",
            "category_subcategory",
            "CategoryValidator.validate_pairs_combined()",
            "Yes",
        ],
        [
            13,
            "Verify Partner Logo on listing page",
            "Listing",
            "Checks partner logo visible on listing",
            "Partner logo displayed",
            "partner_name",
            "listing_validator.validate_partner_logo_on_listing()",
            "Yes",
        ],
        [
            14,
            "Click Quick View and validate product info",
            "Listing",
            "Opens Quick View if available; warns and continues if not",
            "Quick View OK or WARNING then continue",
            "product_name, expected_short_description",
            "validate_quick_view()",
            "Yes",
        ],
        [
            15,
            "Open Product Detail page",
            "Detail",
            "Opens Product Details link for product",
            "Detail page URL /partner-spotlight/... loads",
            "product_name, search_term",
            "open_product_details_link()",
            "Yes",
        ],
        [
            16,
            "Verify page metadata matches product",
            "Detail",
            "Validates title / meta description expectations from Excel",
            "Metadata matches expected values",
            "expected_title, expected_meta_description, expected_keywords",
            "MetadataValidator",
            "Yes",
        ],
        [
            17,
            "Validate Product Name on detail page",
            "Detail",
            "Validates product title on detail",
            "Title matches expected_title / product_name",
            "expected_title, product_name",
            "validate_product_detail()",
            "Yes",
        ],
        [
            18,
            "Verify Thumbnail Image",
            "Detail",
            "Checks product thumbnail image loaded",
            "Image present and loaded",
            "product_name",
            "validate_thumbnail()",
            "Yes",
        ],
        [
            19,
            "Validate Breadcrumb Navigation",
            "Detail",
            "Validates breadcrumb trail present/valid",
            "Breadcrumb OK",
            "(optional) expected_breadcrumb if present",
            "BreadcrumbValidator",
            "Yes",
        ],
        [
            20,
            "Verify complete Product Description",
            "Detail",
            "Validates full description content",
            "Description contains expected text",
            "expected_description",
            "validate_full_description()",
            "Yes",
        ],
        [
            21,
            "Validate Partner Contact link redirect",
            "Detail",
            "Validates contact control supports mailto: and/or partner site URL (including javascript:void(0) + onclick mailto)",
            "All expected_contact_url fragments match",
            "expected_contact_url (e.g. mailto:email; https://...)",
            "detail_validator.validate_partner_contact_link()",
            "Yes",
        ],
        [
            22,
            "Verify Features section",
            "Detail",
            "STRICT: every feature from expected_features must appear on detail page",
            "All features found (one Excel report row per feature)",
            "expected_features (numbered list)",
            "validate_features_section(strict=True)",
            "Yes",
        ],
        [
            23,
            "Validate Resources section links",
            "Detail",
            "Validates resource/download links; asserts if expected_resource_url set",
            "Resource link found / matches URL",
            "expected_resource_url",
            "validate_resource_links()",
            "Yes",
        ],
        [
            24,
            "Download and validate PDF",
            "Detail",
            "Downloads PDF if validate_pdf=TRUE and checks text",
            "PDF downloaded; optional text found",
            "validate_pdf, expected_pdf_text",
            "download_and_validate_pdf()",
            "Yes",
        ],
        [
            25,
            "Verify Categories section",
            "Detail",
            "STRICT combined listing + detail: every category_subcategory pair must be found on site",
            "All Category: Sub Category pairs present",
            "category_subcategory",
            "CategoryValidator.validate_pairs_combined()",
            "Yes",
        ],
        [
            26,
            "Verify Related Products section",
            "Detail",
            "Checks related products; soft WARNING if missing",
            "Related products shown (partner preference)",
            "partner_name",
            "validate_related_products()",
            "Yes",
        ],
    ]
    write_table(
        ws3,
        [
            "Step",
            "Step Title",
            "Area",
            "What Script Does",
            "Expected Result (Pass Criteria)",
            "Excel Data Used",
            "Code / Module",
            "Screenshot",
        ],
        steps,
        [8, 42, 18, 55, 40, 32, 40, 12],
        row_height=55,
    )

    # 4. Test Data Columns
    ws4 = wb.create_sheet("4. Test Data Columns")
    cols = [
        ["enabled", "Yes", "TRUE / FALSE", "Collection", "Only TRUE/1/yes rows are executed"],
        ["test_id", "Yes", "PS-001", "All + reports", "Unique case id for reports/screenshots"],
        ["partner_name", "Yes", "Ecrio", "6, 13, 26", "Partner display name"],
        ["partner_dropdown_label", "Recommended", "Ecrio", "6", "Exact Filter By Partners dropdown text"],
        ["product_name", "Yes", "Ecrio Edge AI Communication Platform", "7–15, 17", "Product / application name on site"],
        ["product_type", "Yes", "application / system", "7", "Edge AI Application or Edge AI System tab preference"],
        ["search_term", "Recommended", "Ecrio Edge AI Communication Platform", "10, 11", "Catalog search term (pSearch); defaults to product_name"],
        ["expected_title", "Recommended", "Ecrio Edge AI Communication Platform", "16, 17", "Detail / metadata title"],
        ["expected_short_description", "Recommended", "Low Latency Edge AI...", "9, 14", "Listing / quick view short text"],
        ["expected_description", "Recommended", "Full description...", "20", "Full product description"],
        ["expected_meta_description", "Optional", "…", "16", "Meta description expectation"],
        ["expected_keywords", "Optional", "…", "16", "Meta keywords"],
        [
            "expected_features",
            "Optional (strict if present)",
            "1. Real-Time Edge AI\n2. On-Device Vision",
            "22",
            "Newline/numbered list; ALL must match detail Features",
        ],
        ["expected_categories", "Optional / legacy", "Vertical: …", "9 tags; fallback", "Legacy; prefer category_subcategory"],
        [
            "category_subcategory",
            "Recommended",
            "Device Type: Edge Device\nVertical: Retail\nUse Cases: …",
            "7, 12, 25",
            "Single ticket column — Category: SubCategory lines",
        ],
        [
            "expected_contact_url",
            "Recommended",
            "mailto:patrick@dynamofl.com; https://partner.com/contact",
            "21",
            "Supports mailto + partner site URL together",
        ],
        ["expected_resource_url", "Optional", "https://www.ecrio.com/", "23", "Expected resource/PDF link fragment"],
        ["validate_pdf", "Optional", "TRUE / FALSE", "24", "If TRUE, download/validate PDF"],
        ["expected_pdf_text", "Optional", "keyword", "24", "Text expected inside downloaded PDF"],
    ]
    write_table(
        ws4,
        ["Column Name", "Required?", "Example", "Used In Steps", "Description"],
        cols,
        [28, 28, 55, 18, 55],
        row_height=40,
    )

    # 5. Catalog Smoke Tests
    ws5 = wb.create_sheet("5. Catalog Smoke Tests")
    smoke = [
        [
            "CAT-001",
            "test_catalog_page_metadata_and_breadcrumbs",
            "smoke, catalog",
            "Open Edge AI Catalog → validate title metadata → validate URL",
            "Title contains Edge AI Catalog; URL matches config",
        ],
        [
            "CAT-002",
            "test_catalog_view_toggle",
            "regression, catalog",
            "Search by Excel search_term → confirm pSearch URL → Grid view → List view",
            "Search returns >0; Grid and/or List active with products AFTER search",
        ],
        [
            "CAT-003",
            "test_catalog_search",
            "smoke, catalog",
            "Search Partner Spotlight; verify URL/content and count",
            "Results > 0; partner-spotlight + search term present",
        ],
        [
            "PS-xxx listing smoke",
            "test_partner_spotlight_listing_smoke",
            "smoke, partner",
            "Open spotlight → search → application name → partner logo",
            "Search + listing name + logo pass",
        ],
    ]
    write_table(
        ws5,
        ["Test ID", "Test Name", "Marker", "Steps", "Pass Criteria"],
        smoke,
        [22, 48, 20, 70, 50],
        row_height=45,
    )

    # 6. Client Ticket Order
    ws6 = wb.create_sheet("6. Client Ticket Order")
    order = [
        [1, "Open site / Engagement / Edge AI Catalog", "1–4", "Menu with direct URL fallback"],
        [2, "Open Partner Spotlight", "4–5", "Explore Partner Spotlight"],
        [3, "Filter By Partners", "6", "Partner dropdown first"],
        [4, "Select Application or System tab", "7", "product_type + resolve if needed"],
        [5, "Apply Category / Sub-Category sidebar filters", "7 (apply) + 12 (report)", "From category_subcategory column"],
        [6, "Verify listing / name / short description", "7–9", "Product card checks"],
        [7, "Search by product name", "10", "Required before Grid/List per UI"],
        [8, "Toggle List View and Grid View", "11", "Only validated on search results page"],
        [9, "Partner logo / Quick View / Product Details", "13–15", "Quick View optional (WARNING)"],
        [10, "Detail page validations", "16–26", "Metadata, content, contact, features, categories, PDF"],
    ]
    write_table(
        ws6,
        ["Order", "Action", "Framework Step", "Notes"],
        order,
        [10, 50, 22, 55],
        row_height=32,
    )

    # 7. Reports
    ws7 = wb.create_sheet("7. Reports")
    reports = [
        [
            "HTML Report",
            r"reports\report.html",
            "pytest-html execution summary",
            "Every run",
        ],
        [
            "Excel Detailed Report",
            r"reports\test_results_YYYYMMDD_HHMMSS.xlsx",
            "4 sheets: Run Summary, Test Summary, Detailed Validations (Expected vs Actual PASS/FAIL/WARNING), QA Recommendations",
            "Every run",
        ],
        [
            "Latest Excel Shortcut",
            r"reports\latest_test_results.xlsx",
            "Copy of most recent Excel report — use this for sharing",
            "Every run",
        ],
        [
            "Allure Results",
            r"reports\allure-results\\",
            "Allure raw results (screenshots + steps)",
            "Every run",
        ],
        [
            "Screenshots",
            r"screenshots\<test_id>\\",
            "Per-step full-page screenshots",
            "Every step when enabled",
        ],
        [
            "Test Log",
            r"reports\test_run_*.log",
            "Console INFO logs",
            "Every run",
        ],
    ]
    write_table(
        ws7,
        ["Report", "Location", "Contents", "When Generated"],
        reports,
        [26, 50, 70, 22],
        row_height=40,
    )

    # 8. Validation Rules
    ws8 = wb.create_sheet("8. Validation Rules")
    rules = [
        ["enabled column", "FALSE/0/blank rows are skipped at collection", "Hard", "Uses is_row_enabled()"],
        [
            "Features (Step 22)",
            "If expected_features filled → ALL features must be found (strict)",
            "Fail test",
            "One Excel report row per feature",
        ],
        [
            "Categories (Step 25)",
            "Listing tags + detail page combined must contain every category_subcategory pair",
            "Fail test",
            "Field name category_subcategory in report",
        ],
        [
            "Partner Contact (Step 21)",
            "Supports mailto: and/or https partner URL; also detects email in onclick/js void(0)",
            "Fail if mismatch",
            "Multiple fragments allowed with ;",
        ],
        [
            "Grid/List (Step 11)",
            "Must run AFTER product search; checks toggle active class (fa-th-large / fa-th-list)",
            "Fail if neither active",
            "Site keeps .listview class even in grid",
        ],
        ["Quick View (Step 14)", "If unavailable → WARNING and continue to Product Details", "Soft", "Not a hard fail"],
        ["Related Products (Step 26)", "Missing section → WARNING", "Soft", "Optional"],
        ["PDF (Step 24)", "Runs only if validate_pdf=TRUE", "Fail if enabled and fails", "Optional"],
        [
            "Excel report safety",
            "Cells starting with =/+/-/@ are sanitized to avoid Excel corruption",
            "Info",
            "excel_sanitize.py",
        ],
    ]
    write_table(
        ws8,
        ["Area", "Rule", "Severity", "Notes"],
        rules,
        [26, 70, 18, 40],
        row_height=40,
    )

    # 9. Project Structure
    ws9 = wb.create_sheet("9. Project Structure")
    struct = [
        ["run_tests.py", "Main entry — pytest runner (visual shortcut)"],
        ["conftest.py", "Browser fixtures, headed/slow-mo, Excel report session finish"],
        ["pytest.ini", "Markers, HTML/Allure options"],
        ["config/config.json", "Timeouts, report paths, testdata path, viewport"],
        ["testdata/partner_products.xlsx", "Excel-driven test data (PartnerProducts sheet)"],
        ["tests/test_partner_spotlight.py", "Parametrized Partner Spotlight tests"],
        ["tests/partner_spotlight_flow.py", "26-step flow implementation"],
        ["tests/test_catalog_features.py", "Catalog smoke/regression"],
        ["pages/*.py", "Page objects (home, catalog, spotlight, detail)"],
        ["utilities/excel_reader.py", "Read Excel + enabled filter"],
        ["utilities/category_parser.py", "Parse category_subcategory ticket format"],
        ["utilities/category_validator.py", "Strict Excel vs site category checks"],
        ["utilities/detail_validator.py", "Detail page validations (features, contact, etc.)"],
        ["utilities/step_reporter.py", "Allure + screenshot + Excel row capture"],
        ["utilities/excel_report_writer.py", "Writes downloadable Expected vs Actual Excel"],
        ["utilities/test_result_store.py", "In-memory validation store for Excel export"],
        ["reports/", "HTML + Excel + logs output (gitignored)"],
        ["screenshots/", "Step screenshots (gitignored)"],
        ["docs/", "This documentation workbook"],
    ]
    write_table(ws9, ["Path", "Role"], struct, [42, 70], row_height=24)

    # 10. QA Tips
    ws10 = wb.create_sheet("10. QA Tips")
    tips = [
        [1, "Keep category_subcategory and expected_features in sync with live site content after every content update.", "High"],
        [2, "Use partner_dropdown_label exactly as shown in Filter By Partners dropdown.", "High"],
        [3, "For contact, put both mailto and site URL in one cell: mailto:a@b.com; https://partner.com/contact", "High"],
        [4, "Validate Grid/List only after search — that matches production UI shown in client screenshots.", "High"],
        [5, "Attach reports/latest_test_results.xlsx to JIRA; use Detailed Validations sheet for Expected vs Actual defect evidence.", "High"],
        [6, "Use python run_tests.py visual for stakeholder demos; use headless for CI/nightly.", "Medium"],
        [7, "Disable WIP rows with enabled=FALSE instead of deleting them.", "Medium"],
        [8, "Quick View / Related Products may be WARNING — decide with PO if they should become hard fails.", "Medium"],
        [9, "Periodic Firefox/WebKit smoke on critical PS-ids for cross-browser confidence.", "Low"],
        [10, "Review Step 7 duration; category filter application can be slow on many pairs.", "Low"],
    ]
    write_table(ws10, ["#", "Recommendation", "Priority"], tips, [6, 95, 12], row_height=32)

    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
