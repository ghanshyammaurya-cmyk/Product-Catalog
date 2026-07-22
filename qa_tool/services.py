"""Run orchestration and safe Excel test-data services for the QA dashboard."""

from __future__ import annotations

import json
import os
import shutil
import signal
import subprocess
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from utilities.excel_sanitize import sanitize_excel_value
from utilities.excel_reader import is_row_enabled

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "qa_tool" / "data"
RUNS_DIR = DATA_DIR / "runs"
REGISTRY_PATH = DATA_DIR / "runs.json"
TEST_DATA_PATH = ROOT / "testdata" / "partner_products.xlsx"
TEST_DATA_SHEET = "PartnerProducts"

ENVIRONMENT_MAP = {"demo": "dev", "qa": "staging", "live": "prod"}
SUITE_ARGS = {
    "full": [],
    "smoke": ["-m", "smoke"],
    "regression": ["-m", "regression"],
    "catalog": ["-m", "catalog"],
    "partner": ["-m", "partner"],
}
BROWSERS = {"chromium", "firefox", "webkit"}
RUN_STATUSES = {"queued", "running", "passed", "failed", "cancelled", "error"}

REQUIRED_COLUMNS = {
    "enabled",
    "test_id",
    "ticket_number",
    "partner_name",
    "partner_dropdown_label",
    "product_name",
    "product_type",
    "search_term",
}
BOOLEAN_COLUMNS = {"enabled", "validate_pdf"}
MULTILINE_COLUMNS = {
    "expected_description",
    "expected_features",
    "expected_categories",
    "category_subcategory",
    "expected_contact_url",
    "expected_resource_url",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _read_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def _atomic_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    os.replace(temp, path)


class TestDataService:
    """Validated, backed-up, atomic access to partner_products.xlsx."""

    def __init__(self, path: Path = TEST_DATA_PATH):
        self.path = path
        self.lock = threading.RLock()

    def read(self) -> dict[str, Any]:
        with self.lock:
            df = pd.read_excel(
                self.path,
                sheet_name=TEST_DATA_SHEET,
                keep_default_na=False,
            )
            rows = []
            for raw in df.to_dict(orient="records"):
                row = {}
                for key, value in raw.items():
                    if key in BOOLEAN_COLUMNS:
                        row[key] = is_row_enabled(value)
                    else:
                        row[key] = "" if value is None else str(value)
                rows.append(row)
            return {
                "columns": list(df.columns),
                "multiline_columns": sorted(MULTILINE_COLUMNS),
                "rows": rows,
                "count": len(rows),
                "enabled_count": sum(bool(row.get("enabled")) for row in rows),
            }

    def write(self, rows: list[dict[str, Any]], run_active: bool = False) -> dict[str, Any]:
        if run_active:
            raise RuntimeError("Test data cannot be edited while a test run is active.")
        if not rows:
            raise ValueError("At least one test-data row is required.")

        with self.lock:
            current = pd.read_excel(
                self.path,
                sheet_name=TEST_DATA_SHEET,
                keep_default_na=False,
            )
            columns = list(current.columns)
            missing_columns = REQUIRED_COLUMNS - set(columns)
            if missing_columns:
                raise ValueError(
                    f"Workbook is missing required columns: {sorted(missing_columns)}"
                )

            normalized = []
            test_ids = []
            for index, raw in enumerate(rows, 1):
                row = {}
                for column in columns:
                    value = raw.get(column, "")
                    if column in BOOLEAN_COLUMNS:
                        row[column] = bool(value) if isinstance(value, bool) else is_row_enabled(value)
                    else:
                        row[column] = sanitize_excel_value(value)

                test_id = str(row.get("test_id", "")).strip()
                partner = str(row.get("partner_name", "")).strip()
                product = str(row.get("product_name", "")).strip()
                if not test_id:
                    raise ValueError(f"Row {index}: test_id is required.")
                if not partner:
                    raise ValueError(f"Row {index}: partner_name is required.")
                if not product:
                    raise ValueError(f"Row {index}: product_name is required.")
                test_ids.append(test_id)
                normalized.append(row)

            duplicates = sorted(
                {test_id for test_id in test_ids if test_ids.count(test_id) > 1}
            )
            if duplicates:
                raise ValueError(f"Duplicate test_id values: {duplicates}")

            backup_dir = self.path.parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = backup_dir / f"partner_products_{stamp}.xlsx"
            shutil.copy2(self.path, backup)

            temp = self.path.with_name(f"{self.path.stem}.tmp.xlsx")
            frame = pd.DataFrame(normalized, columns=columns)
            with pd.ExcelWriter(temp, engine="openpyxl") as writer:
                frame.to_excel(writer, sheet_name=TEST_DATA_SHEET, index=False)

            workbook = load_workbook(temp, read_only=True)
            try:
                if TEST_DATA_SHEET not in workbook.sheetnames:
                    raise ValueError("Generated workbook is missing PartnerProducts sheet.")
            finally:
                workbook.close()

            os.replace(temp, self.path)
            try:
                backup_display = str(backup.relative_to(ROOT))
            except ValueError:
                backup_display = str(backup)
            return {
                "message": "Test data saved successfully.",
                "backup": backup_display,
                "count": len(normalized),
                "enabled_count": sum(bool(row["enabled"]) for row in normalized),
            }


class RunManager:
    """Serial subprocess runner with persistent history and artifact snapshots."""

    def __init__(self):
        RUNS_DIR.mkdir(parents=True, exist_ok=True)
        self.lock = threading.RLock()
        self.runs: dict[str, dict[str, Any]] = {
            item["id"]: item for item in _read_json(REGISTRY_PATH, [])
        }
        self.processes: dict[str, subprocess.Popen] = {}
        # Runs cannot survive a dashboard restart.
        for run in self.runs.values():
            if run.get("status") in ("queued", "running"):
                run["status"] = "error"
                run["finished_at"] = utc_now()
                run["message"] = "Dashboard restarted while run was active."
        self._persist()

    def _persist(self):
        ordered = sorted(
            self.runs.values(), key=lambda item: item.get("created_at", ""), reverse=True
        )
        _atomic_json(REGISTRY_PATH, ordered[:200])

    def active_run(self) -> dict[str, Any] | None:
        with self.lock:
            return next(
                (
                    run
                    for run in self.runs.values()
                    if run.get("status") in ("queued", "running")
                ),
                None,
            )

    def list_runs(self) -> list[dict[str, Any]]:
        with self.lock:
            return sorted(
                (dict(run) for run in self.runs.values()),
                key=lambda item: item.get("created_at", ""),
                reverse=True,
            )

    def get(self, run_id: str) -> dict[str, Any] | None:
        with self.lock:
            run = self.runs.get(run_id)
            return dict(run) if run else None

    def start(self, request: dict[str, Any]) -> dict[str, Any]:
        environment = str(request.get("environment", "qa")).lower()
        suite = str(request.get("suite", "smoke")).lower()
        browser = str(request.get("browser", "chromium")).lower()
        headed = bool(request.get("headed", False))
        slow_mo = int(request.get("slow_mo", 500))
        test_id = str(request.get("test_id", "")).strip()

        if environment not in ENVIRONMENT_MAP:
            raise ValueError("Environment must be demo, qa, or live.")
        if suite not in SUITE_ARGS:
            raise ValueError(f"Suite must be one of: {', '.join(SUITE_ARGS)}")
        if browser not in BROWSERS:
            raise ValueError(f"Browser must be one of: {', '.join(sorted(BROWSERS))}")
        if slow_mo < 0 or slow_mo > 5000:
            raise ValueError("slow_mo must be between 0 and 5000 milliseconds.")
        if test_id and not all(ch.isalnum() or ch in "-_[]" for ch in test_id):
            raise ValueError("test_id contains unsupported characters.")

        with self.lock:
            if self.active_run():
                raise RuntimeError("Another test run is already active.")

            run_id = datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:6]
            run_dir = RUNS_DIR / run_id
            artifact_dir = run_dir / "artifacts"
            artifact_dir.mkdir(parents=True)
            run = {
                "id": run_id,
                "status": "queued",
                "environment": environment,
                "framework_environment": ENVIRONMENT_MAP[environment],
                "suite": suite,
                "browser": browser,
                "headed": headed,
                "slow_mo": slow_mo,
                "test_id": test_id,
                "created_at": utc_now(),
                "started_at": None,
                "finished_at": None,
                "exit_code": None,
                "message": "Queued",
                "artifact_count": 0,
            }
            self.runs[run_id] = run
            self._persist()

            thread = threading.Thread(
                target=self._execute, args=(run_id,), daemon=True, name=f"qa-run-{run_id}"
            )
            thread.start()
            return dict(run)

    def _command(self, run: dict[str, Any], artifact_dir: Path) -> list[str]:
        python = ROOT / "venv" / "Scripts" / "python.exe"
        if not python.exists():
            python = Path(os.sys.executable)
        command = [
            str(python),
            "-u",
            str(ROOT / "run_tests.py"),
            *SUITE_ARGS[run["suite"]],
            f"--env={run['framework_environment']}",
            f"--browser-name={run['browser']}",
            "--reruns",
            "0",
            f"--html={artifact_dir / 'report.html'}",
            f"--alluredir={artifact_dir / 'allure-results'}",
        ]
        if run["headed"]:
            command.extend(["--headed", "--slow-mo", str(run["slow_mo"])])
        if run["test_id"]:
            command.extend(["-k", run["test_id"]])
        return command

    def _execute(self, run_id: str) -> None:
        run_dir = RUNS_DIR / run_id
        artifact_dir = run_dir / "artifacts"
        console_path = run_dir / "console.log"
        automation_log = artifact_dir / "automation.log"

        with self.lock:
            run = self.runs[run_id]
            run["status"] = "running"
            run["started_at"] = utc_now()
            run["message"] = "Test execution in progress"
            self._persist()

        command = self._command(run, artifact_dir)
        process_env = os.environ.copy()
        process_env["PYTHONUNBUFFERED"] = "1"
        process_env["TEST_LOG_FILE"] = str(automation_log)

        creationflags = 0
        if os.name == "nt":
            creationflags = subprocess.CREATE_NEW_PROCESS_GROUP

        try:
            with console_path.open("w", encoding="utf-8", errors="replace") as console:
                console.write("$ " + subprocess.list2cmdline(command) + "\n\n")
                console.flush()
                process = subprocess.Popen(
                    command,
                    cwd=ROOT,
                    env=process_env,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    bufsize=1,
                    shell=False,
                    creationflags=creationflags,
                )
                with self.lock:
                    self.processes[run_id] = process

                assert process.stdout is not None
                for line in process.stdout:
                    console.write(line)
                    console.flush()
                exit_code = process.wait()

            with self.lock:
                run = self.runs[run_id]
                if run.get("status") != "cancelled":
                    run["status"] = "passed" if exit_code == 0 else "failed"
                    run["message"] = (
                        "All selected tests passed."
                        if exit_code == 0
                        else "One or more selected tests failed."
                    )
                run["exit_code"] = exit_code
                run["finished_at"] = utc_now()
        except Exception as exc:
            with self.lock:
                run = self.runs[run_id]
                run["status"] = "error"
                run["message"] = str(exc)
                run["finished_at"] = utc_now()
        finally:
            self._collect_artifacts(run_id)
            with self.lock:
                self.processes.pop(run_id, None)
                self.runs[run_id]["artifact_count"] = len(self.artifacts(run_id))
                self._persist()

    def _collect_artifacts(self, run_id: str):
        artifact_dir = RUNS_DIR / run_id / "artifacts"
        reports = ROOT / "reports"

        # Excel writer still uses the framework report_path. Snapshot only the
        # latest completed Excel into this immutable run directory.
        latest_excel = reports / "latest_test_results.xlsx"
        if latest_excel.exists():
            shutil.copy2(latest_excel, artifact_dir / "test_results.xlsx")

        console = RUNS_DIR / run_id / "console.log"
        if console.exists():
            shutil.copy2(console, artifact_dir / "console.log")

    def cancel(self, run_id: str) -> dict[str, Any]:
        with self.lock:
            run = self.runs.get(run_id)
            if not run:
                raise KeyError("Run not found.")
            if run.get("status") not in ("queued", "running"):
                return dict(run)
            run["status"] = "cancelled"
            run["message"] = "Cancelled by user"
            run["finished_at"] = utc_now()
            process = self.processes.get(run_id)
            self._persist()

        if process and process.poll() is None:
            if os.name == "nt":
                subprocess.run(
                    ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                    capture_output=True,
                    check=False,
                )
            else:
                os.killpg(os.getpgid(process.pid), signal.SIGTERM)
        return self.get(run_id) or {}

    def log(self, run_id: str, offset: int = 0) -> dict[str, Any]:
        path = RUNS_DIR / run_id / "console.log"
        if not path.exists():
            return {"offset": 0, "content": ""}
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            handle.seek(max(0, offset))
            content = handle.read()
            return {"offset": handle.tell(), "content": content}

    def artifacts(self, run_id: str) -> list[dict[str, Any]]:
        artifact_dir = RUNS_DIR / run_id / "artifacts"
        if not artifact_dir.exists():
            return []
        items = []
        for path in sorted(artifact_dir.rglob("*")):
            if not path.is_file():
                continue
            relative = path.relative_to(artifact_dir).as_posix()
            items.append(
                {
                    "name": relative,
                    "size": path.stat().st_size,
                    "url": f"/api/runs/{run_id}/artifacts/{relative}",
                    "viewable": path.suffix.lower() in (".html", ".txt", ".log"),
                }
            )
        return items

    def artifact_path(self, run_id: str, relative: str) -> Path:
        artifact_dir = (RUNS_DIR / run_id / "artifacts").resolve()
        target = (artifact_dir / relative).resolve()
        if artifact_dir not in target.parents or not target.is_file():
            raise FileNotFoundError("Artifact not found.")
        return target


def environment_summary() -> list[dict[str, Any]]:
    output = []
    labels = {"demo": "Demo", "qa": "QA", "live": "Live"}
    colors = {"demo": "purple", "qa": "amber", "live": "green"}
    for public_name, framework_name in ENVIRONMENT_MAP.items():
        path = ROOT / "config" / "environments" / f"{framework_name}.json"
        config = _read_json(path, {})
        warnings = []
        base_url = config.get("base_url", "")
        if public_name == "demo" and ".onsumaye.com" not in base_url:
            warnings.append("Demo URL configuration may be invalid.")
        if public_name == "qa":
            prod = _read_json(
                ROOT / "config" / "environments" / "prod.json", {}
            )
            if base_url and base_url == prod.get("base_url"):
                warnings.append("QA currently targets the same base URL as Live.")
        output.append(
            {
                "id": public_name,
                "label": labels[public_name],
                "framework_name": framework_name,
                "color": colors[public_name],
                "base_url": base_url,
                "edge_catalog_url": config.get("edge_catalog_url", ""),
                "partner_spotlight_url": config.get("partner_spotlight_url", ""),
                "warnings": warnings,
            }
        )
    return output
